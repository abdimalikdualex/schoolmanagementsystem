"""
Admission Officer views - user_type='6'.
Handles student registration, class allocation, documents, and admission reports.
"""
import csv
import io
from datetime import date
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.files.storage import FileSystemStorage
from django.db import transaction
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_http_methods

from .forms import AddStudentForm
from .models import (
    AcademicTerm,
    AdmissionSetting,
    Course,
    CustomUser,
    FeeBalance,
    FeeStructure,
    Guardian,
    Session,
    Student,
    StudentClassEnrollment,
    StudentDocument,
    StudentSubjectEnrollment,
    Subject,
)

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def _get_school(request):
    return getattr(request, 'school', None)


def _create_fee_balance_for_enrollment(student, school_class, session, active_term=None):
    """Create FeeBalance from FeeStructure for class+session."""
    structure = FeeStructure.objects.filter(
        course=school_class,
        session=session,
        is_active=True
    ).select_related('fee_group').first()

    total_fees = Decimal('0')
    due_date = None
    fee_structure = None

    if structure:
        fee_structure = structure
        total_fees = sum(item.amount for item in structure.fee_group.fee_items.all())
        due_date = structure.due_date

    fb, created = FeeBalance.objects.get_or_create(
        student=student,
        session=session,
        defaults={
            'total_fees': total_fees,
            'fee_structure': fee_structure,
            'due_date': due_date,
        }
    )
    if not created and total_fees > 0:
        new_struct_id = fee_structure.id if fee_structure else None
        if fb.total_fees == 0 or (new_struct_id and fb.fee_structure_id != new_struct_id):
            fb.total_fees = total_fees
            fb.fee_structure = fee_structure
            fb.due_date = due_date
            fb.update_balance()
    return fb


@login_required
def admission_dashboard(request):
    """Admission officer dashboard with stats."""
    school = _get_school(request)
    today = timezone.now().date()
    active_term = AcademicTerm.get_active_term(school=school)

    # Base querysets
    students_qs = Student.objects.filter(admin__school=school) if school else Student.objects.all()
    classes_qs = Course.objects.filter(school=school, is_active=True) if school else Course.objects.filter(is_active=True)

    # Stats
    students_today = students_qs.filter(admission_date=today).count()
    students_this_term = 0
    if active_term:
        students_this_term = StudentClassEnrollment.objects.filter(
            academic_year__school=school if school else None,
            term=active_term,
            status='active'
        ).count()
    pending = students_qs.filter(
        Q(status='pending') |
        Q(current_class__isnull=True) |
        ~Q(enrollments__status='active')
    ).distinct().count()

    # Classes with available seats
    classes_with_seats = []
    for cls in classes_qs:
        enrolled = cls.enrollments.filter(status='active').count()
        available = max(0, (cls.capacity or 40) - enrolled)
        classes_with_seats.append({
            'class': cls,
            'enrolled': enrolled,
            'capacity': cls.capacity or 40,
            'available': available,
        })

    context = {
        'page_title': 'Admission Dashboard',
        'students_today': students_today,
        'students_this_term': students_this_term,
        'pending': pending,
        'classes_with_seats': classes_with_seats,
        'active_term': active_term,
    }
    return render(request, 'admission_template/admission_dashboard.html', context)


@login_required
def new_student_admission(request):
    """New student admission - reuses add_student logic with fee loading."""
    school = _get_school(request)
    form = AddStudentForm(request.POST or None, request.FILES or None, school=school)
    active_term = AcademicTerm.get_active_term(school=school)
    sessions = Session.objects.filter(school=school) if school else Session.objects.all()

    # Fee preview when class selected (AJAX)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and request.method == 'GET':
        class_id = request.GET.get('class_id')
        session_id = request.GET.get('session_id')
        if not session_id and sessions.exists():
            session_id = sessions.first().id
        if class_id and session_id:
            try:
                cls = Course.objects.get(id=class_id, school=school)
                session = Session.objects.get(id=session_id, school=school)
                structure = FeeStructure.objects.filter(
                    course=cls,
                    session=session,
                    is_active=True
                ).select_related('fee_group').first()
                total = Decimal('0')
                if structure and structure.fee_group:
                    total = sum(item.amount for item in structure.fee_group.fee_items.all())
                return JsonResponse({'total_fee': str(total), 'currency': 'KES'})
            except (Course.DoesNotExist, Session.DoesNotExist):
                pass
        return JsonResponse({'total_fee': '0', 'currency': 'KES'})

    context = {
        'form': form,
        'page_title': 'New Student Admission',
        'active_term': active_term,
        'sessions': sessions,
    }

    if request.method == 'POST':
        if school and not school.can_add_student():
            plan = school.subscription_plan
            limit_str = "Unlimited" if plan.student_limit == 0 else str(plan.student_limit)
            messages.error(request, f"Student limit reached ({limit_str}). Upgrade your plan.")
            return render(request, 'admission_template/admission_form.html', context)

        if form.is_valid():
            first_name = form.cleaned_data['first_name'].strip()
            last_name = form.cleaned_data['last_name'].strip()
            email = form.cleaned_data['email'].strip().lower()
            gender = form.cleaned_data['gender']
            password = form.cleaned_data['password']
            address = form.cleaned_data.get('address') or 'N/A'
            course = form.cleaned_data['course']
            guardian_name = form.cleaned_data['guardian_name'].strip()
            guardian_email = form.cleaned_data.get('guardian_email') or ''
            guardian_phone = form.cleaned_data['guardian_phone'].strip()
            passport = request.FILES.get('profile_pic')
            admission_number = form.cleaned_data['admission_number'].strip()

            if not active_term:
                messages.error(request, "No active term. Please activate an academic term first.")
                return render(request, 'admission_template/admission_form.html', context)

            try:
                passport_url = ''
                if passport:
                    fs = FileSystemStorage()
                    filename = fs.save(passport.name, passport)
                    passport_url = fs.url(filename)

                with transaction.atomic():
                    user = CustomUser.objects.create_user(
                        email=email,
                        password=password,
                        user_type='3',
                        first_name=first_name,
                        last_name=last_name,
                        profile_pic=passport_url,
                    )
                    user.gender = gender
                    user.address = address
                    user.phone_number = form.cleaned_data.get('phone_number') or ''
                    user.school = school
                    user.save()

                    student = Student.objects.get(admin=user)
                    student.admission_number = admission_number
                    student.course = course
                    student.current_class = course
                    student.status = 'active'
                    student.admission_date = timezone.now().date()
                    student.save()

                    Guardian.objects.create(
                        student=student,
                        name=guardian_name,
                        email=guardian_email or None,
                        phone_number=guardian_phone,
                        is_primary=True,
                    )

                    session = Session.objects.filter(
                        school=school,
                        academic_year=timezone.now().year
                    ).first() or Session.objects.filter(school=school).first()
                    if session:
                        student.session = session
                        student.save()

                    if not session:
                        raise ValueError("No Session found. Please add an academic session first.")

                    enrollment = StudentClassEnrollment.objects.create(
                        student=student,
                        school_class=course,
                        academic_year=session,
                        term=active_term,
                        status='active',
                    )

                    subjects = Subject.objects.filter(course=course).filter(
                        Q(term=active_term) | Q(term__isnull=True)
                    )
                    for subj in subjects:
                        StudentSubjectEnrollment.objects.get_or_create(
                            student=student,
                            subject=subj,
                            term=active_term,
                            defaults={'enrollment': enrollment},
                        )

                    _create_fee_balance_for_enrollment(student, course, session, active_term)

                messages.success(request, f"Student {first_name} {last_name} added. Admission No: {admission_number}")
                return redirect(reverse('admission_dashboard'))
            except Exception as e:
                messages.error(request, "Could not add student: " + str(e))
        else:
            messages.error(request, "Please fulfil all requirements")

    return render(request, 'admission_template/admission_form.html', context)


@login_required
def bulk_admission(request):
    """Bulk admission from Excel/CSV."""
    school = _get_school(request)
    active_term = AcademicTerm.get_active_term(school=school)
    sessions = Session.objects.filter(school=school) if school else Session.objects.all()
    classes_qs = Course.objects.filter(school=school, is_active=True) if school else Course.objects.none()

    if request.method == 'POST':
        file = request.FILES.get('file')
        session_id = request.POST.get('session_id')
        class_id = request.POST.get('class_id')

        if not file or not session_id or not class_id:
            messages.error(request, "Please select file, session, and class.")
            return redirect('admission_bulk')

        session = get_object_or_404(Session, id=session_id, school=school)
        school_class = get_object_or_404(Course, id=class_id, school=school)

        if not active_term:
            messages.error(request, "No active term.")
            return redirect('admission_bulk')

        # Capacity check
        enrolled = school_class.enrollments.filter(status='active').count()
        capacity = school_class.capacity or 40

        rows = []
        filename = file.name.lower()
        try:
            if filename.endswith('.csv'):
                content = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)
            elif filename.endswith(('.xlsx', '.xls')) and OPENPYXL_AVAILABLE:
                wb = openpyxl.load_workbook(file, read_only=True)
                ws = wb.active
                headers = [str(c.value or '').strip() for c in ws[1]]
                for row in ws.iter_rows(min_row=2):
                    vals = [str(c.value or '').strip() if c.value is not None else '' for c in row]
                    rows.append(dict(zip(headers, vals)))
                wb.close()
            else:
                messages.error(request, "Upload CSV or Excel (.xlsx). Install openpyxl for Excel support.")
                return redirect('admission_bulk')
        except Exception as e:
            messages.error(request, f"Could not read file: {str(e)}")
            return redirect('admission_bulk')

        # Expected columns: first_name, last_name, email, gender, guardian_name, guardian_phone
        created = 0
        skipped = 0
        errors = []
        today = timezone.now().date()

        for i, row in enumerate(rows):
            if enrolled >= capacity:
                errors.append(f"Row {i+2}: Class capacity reached.")
                skipped += 1
                continue

            first_name = (row.get('first_name') or row.get('First Name') or '').strip()
            last_name = (row.get('last_name') or row.get('Last Name') or '').strip()
            email = (row.get('email') or row.get('Email') or '').strip().lower()
            gender_raw = (row.get('gender') or row.get('Gender') or 'M').strip().upper()
            gender = 'F' if gender_raw.startswith('F') else 'M'
            guardian_name = (row.get('guardian_name') or row.get('Guardian Name') or '').strip()
            guardian_phone = (row.get('guardian_phone') or row.get('Guardian Phone') or '').strip()

            if not first_name or not last_name or not email:
                errors.append(f"Row {i+2}: Missing first_name, last_name, or email.")
                skipped += 1
                continue

            if CustomUser.objects.filter(email=email).exists():
                errors.append(f"Row {i+2}: Email {email} already exists.")
                skipped += 1
                continue

            try:
                with transaction.atomic():
                    setting = AdmissionSetting.objects.filter(school=school).first()
                    if not setting:
                        setting = AdmissionSetting.objects.create(
                            prefix='ADM', start_number=1000, next_number=1000, school=school
                        )
                    admission_number = setting.get_next_admission()

                    user = CustomUser.objects.create_user(
                        email=email,
                        password='changeme123',
                        user_type='3',
                        first_name=first_name,
                        last_name=last_name,
                        gender=gender,
                        address='N/A',
                        school=school,
                    )
                    student = Student.objects.get(admin=user)
                    student.admission_number = admission_number
                    student.current_class = school_class
                    student.course = school_class
                    student.session = session
                    student.status = 'active'
                    student.admission_date = today
                    student.save()

                    Guardian.objects.create(
                        student=student,
                        name=guardian_name or 'Guardian',
                        phone_number=guardian_phone or 'N/A',
                        is_primary=True,
                    )

                    enrollment = StudentClassEnrollment.objects.create(
                        student=student,
                        school_class=school_class,
                        academic_year=session,
                        term=active_term,
                        status='active',
                    )

                    _create_fee_balance_for_enrollment(student, school_class, session, active_term)

                    created += 1
                    enrolled += 1
            except Exception as e:
                errors.append(f"Row {i+2}: {str(e)}")
                skipped += 1

        if errors:
            for err in errors[:10]:
                messages.warning(request, err)
            if len(errors) > 10:
                messages.warning(request, f"... and {len(errors)-10} more errors.")
        messages.success(request, f"Bulk admission: {created} created, {skipped} skipped.")
        return redirect('admission_dashboard')

    context = {
        'page_title': 'Bulk Admission',
        'active_term': active_term,
        'sessions': sessions,
        'classes': classes_qs,
        'openpyxl_available': OPENPYXL_AVAILABLE,
    }
    return render(request, 'admission_template/admission_bulk.html', context)


@login_required
def class_allocation(request):
    """Assign students to classes with capacity checks."""
    school = _get_school(request)
    active_term = AcademicTerm.get_active_term(school=school)
    sessions = Session.objects.filter(school=school) if school else Session.objects.all()
    classes_qs = Course.objects.filter(school=school, is_active=True) if school else Course.objects.none()

    # Unallocated students (no class or no active enrollment this term)
    unallocated = Student.objects.none()
    if school:
        enrolled_ids = StudentClassEnrollment.objects.filter(
            academic_year__school=school,
            term=active_term,
            status='active'
        ).values_list('student_id', flat=True)
        unallocated = Student.objects.filter(
            admin__school=school
        ).exclude(id__in=enrolled_ids).filter(
            Q(status='active') | Q(status='pending')
        ).select_related('admin', 'current_class')[:200]

    if request.method == 'POST':
        student_ids = request.POST.getlist('student_ids')
        class_id = request.POST.get('class_id')
        session_id = request.POST.get('session_id')

        if not student_ids or not class_id or not session_id:
            messages.error(request, "Select students, class, and session.")
            return redirect('admission_class_allocation')

        school_class = get_object_or_404(Course, id=class_id, school=school)
        session = get_object_or_404(Session, id=session_id, school=school)
        capacity = school_class.capacity or 40
        current_enrollment = school_class.enrollments.filter(status='active').count()
        available = max(0, capacity - current_enrollment)

        allocated = 0
        for sid in student_ids[:available]:
            student = Student.objects.filter(id=sid, admin__school=school).first()
            if not student:
                continue
            if student.enrollments.filter(academic_year=session, status='active').exists():
                continue
            try:
                with transaction.atomic():
                    student.current_class = school_class
                    student.course = school_class
                    student.session = session
                    student.status = 'active'
                    student.save()

                    enrollment = StudentClassEnrollment.objects.create(
                        student=student,
                        school_class=school_class,
                        academic_year=session,
                        term=active_term,
                        status='active',
                    )

                    _create_fee_balance_for_enrollment(student, school_class, session, active_term)
                    allocated += 1
            except Exception:
                pass

        messages.success(request, f"Allocated {allocated} students to {school_class.name}.")
        return redirect('admission_class_allocation')

    # Classes with capacity info
    classes_with_seats = []
    for cls in classes_qs:
        enrolled = cls.enrollments.filter(status='active').count()
        classes_with_seats.append({
            'class': cls,
            'enrolled': enrolled,
            'capacity': cls.capacity or 40,
            'available': max(0, (cls.capacity or 40) - enrolled),
        })

    context = {
        'page_title': 'Class Allocation',
        'unallocated': unallocated,
        'classes_with_seats': classes_with_seats,
        'sessions': sessions,
        'active_term': active_term,
    }
    return render(request, 'admission_template/admission_class_allocation.html', context)


@login_required
def student_documents(request, student_id=None):
    """Upload/manage student documents."""
    school = _get_school(request)
    students_qs = Student.objects.filter(admin__school=school) if school else Student.objects.none()

    if student_id:
        student = get_object_or_404(Student, id=student_id, admin__school=school)
        documents = student.documents.all()

        if request.method == 'POST':
            doc_type = request.POST.get('document_type', 'other')
            file = request.FILES.get('file')
            notes = request.POST.get('notes', '').strip()
            if file:
                doc = StudentDocument.objects.create(
                    student=student,
                    document_type=doc_type,
                    file=file,
                    notes=notes or None,
                )
                messages.success(request, "Document uploaded.")
                return redirect('admission_student_documents', student_id=student.id)
            else:
                messages.error(request, "Please select a file.")

        context = {
            'page_title': f'Documents - {student}',
            'student': student,
            'documents': documents,
        }
        return render(request, 'admission_template/admission_documents.html', context)

    # List students
    context = {
        'page_title': 'Student Documents',
        'students': students_qs.select_related('admin', 'current_class')[:100],
    }
    return render(request, 'admission_template/admission_documents_list.html', context)


@login_required
def admission_document_delete(request, document_id):
    """Delete a student document."""
    school = _get_school(request)
    doc = get_object_or_404(StudentDocument, id=document_id, student__admin__school=school)
    student_id = doc.student_id
    doc.delete()
    messages.success(request, "Document deleted.")
    return redirect('admission_student_documents', student_id=student_id)


@login_required
def admission_reports(request):
    """Admission reports by class, gender, term."""
    school = _get_school(request)
    session_id = request.GET.get('session_id')
    sessions = Session.objects.filter(school=school) if school else Session.objects.all()
    session = sessions.filter(id=session_id).first() if session_id else sessions.order_by('-start_year').first()

    by_class = []
    by_gender = []
    by_term = []

    if session:
        enrollments = StudentClassEnrollment.objects.filter(
            academic_year=session,
            status='active'
        ).select_related('student', 'school_class', 'student__admin')

        by_class = list(
            enrollments.values('school_class__name')
            .annotate(count=Count('id'))
            .order_by('-count')
        )

        student_ids = list(enrollments.values_list('student_id', flat=True).distinct())
        by_gender = list(
            Student.objects.filter(id__in=student_ids)
            .values('admin__gender')
            .annotate(count=Count('id'))
        )

    context = {
        'page_title': 'Admission Reports',
        'session': session,
        'sessions': sessions,
        'by_class': by_class,
        'by_gender': by_gender,
    }
    return render(request, 'admission_template/admission_reports.html', context)
